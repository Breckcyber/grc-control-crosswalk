#!/usr/bin/env python3
"""
export_excel.py
===============
Generate a professional, multi-sheet Excel audit workbook for the
GRC Control Crosswalk project.

The workbook is built from three YAML data files in the repo's data/ folder:

    data/mappings.yaml         -> Sheet 1 "Control Mapping"
    data/audit-readiness.yaml  -> Sheet 2 "Audit Evidence" and Sheet 3 "Control Testing"
    data/risk-register.yaml    -> Sheet 4 "Risk Register" and Sheet 5 "Remediation Planning"

NOTE FOR THE ANALYST:
The five framework files (nist-csf.yaml, iso-27001.yaml, etc.) are NOT read here.
Every value these five sheets need already lives in the three files above —
including control names, which are stored inside audit-readiness.yaml. If a
future sheet needs to pull control titles or descriptions directly from a
framework file, add a loader for it following the same pattern used below.

Output:  GRC_Audit_Workbook.xlsx  (written to the repo root)

Run it from anywhere — paths are resolved relative to THIS script's location,
not the current working directory:

    python src/export_excel.py
    python /full/path/to/repo/src/export_excel.py

Requires:  openpyxl, pyyaml   (pip install openpyxl pyyaml)
"""

import sys
from datetime import date
from math import ceil
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. PATHS  (resolved relative to this file so the script is location-proof)
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent      # .../repo/src
REPO_ROOT = SCRIPT_DIR.parent                     # .../repo
DATA_DIR = REPO_ROOT / "data"                     # .../repo/data
OUTPUT_FILE = REPO_ROOT / "GRC_Audit_Workbook.xlsx"

TODAY = date.today().strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# 2. STYLING CONSTANTS  (one place to tweak the look of the whole workbook)
# ---------------------------------------------------------------------------
NAVY = "1F3864"          # brand navy, used for the title + header banners
WHITE = "FFFFFF"

# Title row (row 1) and header row (row 2) share the navy banner look.
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BANNER_FILL = PatternFill(fill_type="solid", start_color=NAVY, end_color=NAVY)

# Body cells.
BODY_FONT = Font(name="Calibri", size=10, color="000000")
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER_TOP = Alignment(wrap_text=True, vertical="top", horizontal="center")

# Light grid so the table reads cleanly.
THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(bottom=Side(style="medium", color="0F2147"))

# Risk Band / Risk Rating colour scale.
# Each entry is (fill_hex, font_color_hex, bold?). Critical is the only inverted
# (dark fill, white text) treatment, per the spec.
BAND_STYLES = {
    "Low":      ("C6EFCE", "006100", False),   # light green
    "Medium":   ("FFE0B2", "7F4F00", False),   # light orange
    "High":     ("FFC7CE", "9C0006", False),   # light red
    "Critical": ("C00000", WHITE,    True),    # dark red, white text
}

# Remediation Status subtle shading.
STATUS_STYLES = {
    "Open":        ("F2F2F2", "404040", False),  # light gray
    "In Progress": ("DDEBF7", "1F4E79", False),  # light blue
    "Mitigated":   ("E2EFDA", "375623", False),  # light green
}


# ---------------------------------------------------------------------------
# 3. SMALL HELPERS
# ---------------------------------------------------------------------------
def load_yaml(filename):
    """Load one YAML file from the data/ folder using yaml.safe_load().

    Exits with a clear, actionable message if the file is missing or invalid,
    rather than dumping a raw traceback on the analyst.
    """
    path = DATA_DIR / filename
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return yaml.safe_load(fh)
    except FileNotFoundError:
        sys.exit(
            f"ERROR: Could not find required data file:\n"
            f"    {path}\n"
            f"Make sure '{filename}' exists in the data/ folder at the repo root."
        )
    except yaml.YAMLError as exc:
        sys.exit(f"ERROR: '{filename}' is not valid YAML.\nDetails: {exc}")


def extract_list(data, *candidate_keys):
    """Return the list of records inside a loaded YAML document.

    The framework/data files wrap their records under a top-level key
    (e.g. 'mappings:', 'controls:', 'risks:') alongside metadata such as
    'version' and 'methodology'. This pulls out the right list whether the
    document is that wrapped dict or a bare top-level list.
    """
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in candidate_keys:
            value = data.get(key)
            if isinstance(value, list):
                return value
    # Nothing matched -> treat as empty so the sheet is created but blank.
    return []


def numbered(items):
    """Turn a list into a single numbered string for one cell:
        ["a", "b"] -> "1. a\\n2. b"
    Returns "" for an empty/missing list.
    """
    if not items:
        return ""
    return "\n".join(f"{i}. {text}" for i, text in enumerate(items, start=1))


def safe(value):
    """Render any scalar as a clean string ('' for None)."""
    return "" if value is None else str(value)


def estimate_row_height(values, widths):
    """Estimate a readable row height (in points) for wrapped text.

    For each cell we work out how many display lines its text will occupy at
    that column's width, honouring any explicit '\\n' line breaks, then take
    the tallest cell in the row. Roughly 15 pts per line, capped so a very long
    cell never produces an absurd row.
    """
    max_lines = 1
    for text, width in zip(values, widths):
        text = safe(text)
        chars_per_line = max(1, width - 1)          # width unit ≈ characters
        cell_lines = 0
        for segment in text.split("\n"):            # respect manual line breaks
            cell_lines += max(1, ceil(len(segment) / chars_per_line))
        max_lines = max(max_lines, cell_lines)
    return min(max_lines * 15, 320)                 # 320 pt hard cap


# ---------------------------------------------------------------------------
# 4. SHEET SCAFFOLDING  (title banner + header banner + widths + freeze)
# ---------------------------------------------------------------------------
def start_sheet(wb, sheet_name, headers, col_widths):
    """Create a sheet and lay down the title row, header row, column widths,
    and frozen panes. Returns the worksheet, ready for data starting at row 3.
    """
    ws = wb.create_sheet(title=sheet_name)
    ncols = len(headers)

    # --- Row 1: title banner spanning all columns ------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(
        row=1, column=1,
        value=f"{sheet_name}    —    GRC Control Crosswalk — generated {TODAY}",
    )
    title_cell.font = TITLE_FONT
    title_cell.fill = BANNER_FILL
    title_cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 26

    # --- Row 2: column headers (bold white on navy) ----------------------
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = BANNER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = HEADER_BORDER
    ws.row_dimensions[2].height = 22

    # --- Column widths ---------------------------------------------------
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths[header]

    # --- Freeze title + header so they stay visible while scrolling ------
    ws.freeze_panes = "A3"
    return ws


def write_row(ws, row_idx, values, widths, *, band_cols=(), status_cols=(), center_cols=()):
    """Write one data row at row_idx with wrapping, borders, row height, and
    optional conditional colouring.

    band_cols / status_cols / center_cols are sets of 0-based COLUMN INDEXES.
      - band_cols   -> apply the Risk Band / Risk Rating colour scale
      - status_cols -> apply the Remediation Status shading
      - center_cols -> horizontally centre the value (e.g. IDs, scores)
    """
    for col_idx, value in enumerate(values):
        cell = ws.cell(row=row_idx, column=col_idx + 1, value=safe(value))
        cell.font = BODY_FONT
        cell.border = CELL_BORDER
        cell.alignment = CENTER_TOP if col_idx in center_cols else WRAP_TOP

        # Risk Band / Risk Rating colour coding.
        if col_idx in band_cols:
            style = BAND_STYLES.get(safe(value))
            if style:
                fill_hex, font_hex, bold = style
                cell.fill = PatternFill(fill_type="solid", start_color=fill_hex, end_color=fill_hex)
                cell.font = Font(name="Calibri", size=10, color=font_hex, bold=bold)
                cell.alignment = CENTER_TOP

        # Remediation Status shading.
        if col_idx in status_cols:
            style = STATUS_STYLES.get(safe(value))
            if style:
                fill_hex, font_hex, bold = style
                cell.fill = PatternFill(fill_type="solid", start_color=fill_hex, end_color=fill_hex)
                cell.font = Font(name="Calibri", size=10, color=font_hex, bold=bold)
                cell.alignment = CENTER_TOP

    # Set a readable height based on the widest-wrapping cell in this row.
    width_list = [ws.column_dimensions[get_column_letter(c + 1)].width for c in range(len(values))]
    ws.row_dimensions[row_idx].height = estimate_row_height(values, width_list)


# ---------------------------------------------------------------------------
# 5. SHEET BUILDERS  (one function per sheet — read top to bottom)
# ---------------------------------------------------------------------------
def build_control_mapping(wb, mappings_doc):
    """SHEET 1 — Control Mapping. One row per mapping TARGET (flattened)."""
    headers = ["Source Framework", "Source ID", "Target Framework",
               "Target ID", "Relationship", "Notes"]
    widths = {"Source Framework": 18, "Source ID": 14, "Target Framework": 18,
              "Target ID": 14, "Relationship": 14, "Notes": 60}
    ws = start_sheet(wb, "Control Mapping", headers, widths)

    mappings = extract_list(mappings_doc, "mappings")
    row = 3
    for entry in mappings:
        source = entry.get("source", {}) or {}
        for target in entry.get("targets", []) or []:
            write_row(
                ws, row,
                [
                    source.get("framework"),
                    source.get("id"),
                    target.get("framework"),
                    target.get("id"),
                    target.get("relationship"),
                    target.get("notes"),
                ],
                widths.values(),
                center_cols={1, 3, 4},   # the ID + framework-target columns
            )
            row += 1
    return row - 3   # number of data rows written


def build_audit_evidence(wb, audit_doc):
    """SHEET 2 — Audit Evidence. Evidence requirements numbered in one cell."""
    headers = ["Framework", "Control ID", "Control Name", "Evidence Requirements"]
    widths = {"Framework": 18, "Control ID": 12, "Control Name": 32, "Evidence Requirements": 70}
    ws = start_sheet(wb, "Audit Evidence", headers, widths)

    controls = extract_list(audit_doc, "controls")
    row = 3
    for ctrl in controls:
        write_row(
            ws, row,
            [
                ctrl.get("framework"),
                ctrl.get("id"),
                ctrl.get("name"),
                numbered(ctrl.get("evidence_requirements")),
            ],
            widths.values(),
            center_cols={1},
        )
        row += 1
    return row - 3


def build_control_testing(wb, audit_doc):
    """SHEET 3 — Control Testing. Testing procedure numbered; Risk Rating
    colour-coded on the same scale as Risk Band (shared High/Medium/... values).
    """
    headers = ["Framework", "Control ID", "Control Name", "Testing Procedure",
               "Pass Criteria", "Fail Criteria", "Sample Finding", "Risk Rating"]
    widths = {"Framework": 16, "Control ID": 11, "Control Name": 26,
              "Testing Procedure": 55, "Pass Criteria": 40, "Fail Criteria": 40,
              "Sample Finding": 45, "Risk Rating": 12}
    ws = start_sheet(wb, "Control Testing", headers, widths)

    controls = extract_list(audit_doc, "controls")
    row = 3
    for ctrl in controls:
        write_row(
            ws, row,
            [
                ctrl.get("framework"),
                ctrl.get("id"),
                ctrl.get("name"),
                numbered(ctrl.get("testing_procedure")),
                ctrl.get("pass_criteria"),
                ctrl.get("fail_criteria"),
                ctrl.get("sample_finding"),
                ctrl.get("risk_rating"),
            ],
            widths.values(),
            band_cols={7},       # Risk Rating column
            center_cols={1},
        )
        row += 1
    return row - 3


def build_risk_register(wb, risk_doc):
    """SHEET 4 — Risk Register. Likelihood/Impact combine number + label
    ('3 (Possible)'); Risk Band colour-coded.
    """
    headers = ["Risk ID", "Framework", "Control ID", "Risk Description",
               "Likelihood", "Impact", "Risk Score", "Risk Band", "Risk Owner"]
    widths = {"Risk ID": 9, "Framework": 16, "Control ID": 11, "Risk Description": 60,
              "Likelihood": 16, "Impact": 16, "Risk Score": 11, "Risk Band": 12,
              "Risk Owner": 32}
    ws = start_sheet(wb, "Risk Register", headers, widths)

    risks = extract_list(risk_doc, "risks")
    row = 3
    for risk in risks:
        ctrl = risk.get("related_control", {}) or {}
        likelihood = f'{safe(risk.get("likelihood"))} ({safe(risk.get("likelihood_label"))})'
        impact = f'{safe(risk.get("impact"))} ({safe(risk.get("impact_label"))})'
        write_row(
            ws, row,
            [
                risk.get("risk_id"),
                ctrl.get("framework"),
                ctrl.get("id"),
                risk.get("risk_description"),
                likelihood,
                impact,
                risk.get("risk_score"),
                risk.get("risk_band"),
                risk.get("risk_owner"),
            ],
            widths.values(),
            band_cols={7},                       # Risk Band column
            center_cols={0, 2, 4, 5, 6},          # IDs, scores, scales
        )
        row += 1
    return row - 3


def build_remediation_planning(wb, risk_doc):
    """SHEET 5 — Remediation Planning. Risk Band colour-coded; Status shaded."""
    headers = ["Risk ID", "Framework", "Control ID", "Risk Band",
               "Risk Owner", "Target Date", "Status", "Treatment"]
    widths = {"Risk ID": 9, "Framework": 16, "Control ID": 11, "Risk Band": 12,
              "Risk Owner": 30, "Target Date": 14, "Status": 14, "Treatment": 60}
    ws = start_sheet(wb, "Remediation Planning", headers, widths)

    risks = extract_list(risk_doc, "risks")
    row = 3
    for risk in risks:
        ctrl = risk.get("related_control", {}) or {}
        write_row(
            ws, row,
            [
                risk.get("risk_id"),
                ctrl.get("framework"),
                ctrl.get("id"),
                risk.get("risk_band"),
                risk.get("risk_owner"),
                risk.get("target_remediation_date"),
                risk.get("remediation_status"),
                risk.get("treatment"),
            ],
            widths.values(),
            band_cols={3},        # Risk Band column
            status_cols={6},      # Status column
            center_cols={0, 2, 5},
        )
        row += 1
    return row - 3


# ---------------------------------------------------------------------------
# 6. MAIN
# ---------------------------------------------------------------------------
def main():
    # Load the three source files (clear error + exit if any are missing).
    mappings_doc = load_yaml("mappings.yaml")
    audit_doc = load_yaml("audit-readiness.yaml")
    risk_doc = load_yaml("risk-register.yaml")

    # Fresh workbook; remove the default empty sheet openpyxl creates.
    wb = Workbook()
    wb.remove(wb.active)

    # Build all five sheets, capturing the data-row count of each.
    counts = {
        "Control Mapping":      build_control_mapping(wb, mappings_doc),
        "Audit Evidence":       build_audit_evidence(wb, audit_doc),
        "Control Testing":      build_control_testing(wb, audit_doc),
        "Risk Register":        build_risk_register(wb, risk_doc),
        "Remediation Planning": build_remediation_planning(wb, risk_doc),
    }

    wb.save(OUTPUT_FILE)

    # Confirmation summary.
    print(f"Created workbook: {OUTPUT_FILE}")
    print("Rows per sheet (excluding title + header rows):")
    for sheet_name, n in counts.items():
        print(f"  - {sheet_name:<22} {n} rows")
    print(f"Total data rows: {sum(counts.values())}")


if __name__ == "__main__":
    main()